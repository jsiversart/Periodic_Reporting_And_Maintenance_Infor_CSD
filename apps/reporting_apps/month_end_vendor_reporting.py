# month end vendor reporting.py

def create_vendor_reports():
    
    import sys
    from pathlib import Path

    # --- Ensure repo root is in Python path ---
    repo_root = Path(__file__).resolve().parent
    while not (repo_root / "core").exists() and repo_root.parent != repo_root:
        repo_root = repo_root.parent

    sys.path.insert(0, str(repo_root))

    import pandas as pd
    import jaydebeapi
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from core.config import PATHS, JDBC, COMPANY_NAME


    # Connect to DB
    conn_remote = jaydebeapi.connect(
        JDBC["class"],
        JDBC["url"],
        [JDBC["user"], JDBC["password"]],
        JDBC["jar"]
    )

    export_folder = PATHS["reports"]

    # -------------------------
    #  DATE LOGIC
    # -------------------------
    today = datetime.today()
    first_of_this_month = today.replace(day=1)
    prev_month_date = first_of_this_month - timedelta(days=1)
    prev_month_yyyymm = prev_month_date.strftime("%Y%m") #YYYYMM

    # -------------------------
    #  QUERIES
    # -------------------------
    queries = {
        f"DL_SUPPLY_POS_{prev_month_yyyymm}": f"""
            WITH RankedRows AS (SELECT
            prod, arpvendno, baseprice, ROW_NUMBER() OVER (PARTITION BY prod ORDER BY CASE WHEN whse = '25' THEN 0 ELSE 1 END) AS RowNum
            FROM default.icsw WHERE cono = '1' AND arpvendno = 640)
            SELECT
            {COMPANY_NAME} AS DistributorName,
            d.whse AS DistributorLocationNumber,
            d.name + ' Branch' AS DistributorWarehouseName,
            d.city AS DistributorWarehouseCity,
            d.statecd AS DistributorWarehouseState,
            CASE 
                WHEN LENGTH(d.zipcd) > 5 THEN SUBSTR(d.zipcd, 1, 5)
                ELSE d.zipcd
                END AS DistributorWarehousePostalCode,
            'US' AS DistributorWarehouseCountry,
            l.shipprod AS DistributorSKUNumber,
            '' AS DistributorUnitCost,
            '' AS DistributorTotalAmount,
            '' AS ContractorLoyaltyProgramNumber,
            '' AS ContractorName,
            '' AS ContractorAddress,
            '' AS ContractorCity,
            '' AS ContractorState,
            h.shiptozip AS ContractorPostalCode,
            '' AS ContractorCountry,
            CONCAT(CAST(l.orderno AS VARCHAR(10)), '-', CAST(l.ordersuf AS VARCHAR(5))) AS InvoiceNumber,
            l.invoicedt AS InvoiceDate,
            '' AS ResideoSKUNumber,
            CASE WHEN upper(l.transtype) = 'RM' THEN -l.qtyship 
                    WHEN l.returnfl = 'True' THEN -l.qtyship 
                    ELSE l.qtyship END AS QuantitySold,
            r.baseprice AS ContractorUnitPrice,
            '' AS ContractorTotalAmount,
            '' AS CurrencyBeingReported
            FROM default.oeel l
            LEFT JOIN oeeh h
            ON h.cono = 1
            AND h.orderno = l.orderno
            AND h.ordersuf = l.ordersuf
            LEFT JOIN icsd d
            ON d.cono = 1
            AND d.whse = l.whse
            LEFT JOIN RankedRows r 
                ON r.prod = l.shipprod
                AND r.RowNum = 1
            WHERE l.cono = 1
            AND l.whse NOT LIKE '%C'
            AND l.canceldt IS NULL
            AND l.shipprod IN (
                    SELECT prod
                    FROM default.icsw
                    WHERE cono = 1
                    AND arpvendno = 640
                )
            AND l.qtyship <> 0
            AND LEFT(LTRIM(RTRIM(l.invoicedt)), 7) =
                CONCAT(
                    CAST(DATEPART(YEAR, DATEADD(MONTH, -1, GETDATE())) AS VARCHAR(4)),
                    '-',
                    RIGHT(CONCAT('0', CAST(DATEPART(MONTH, DATEADD(MONTH, -1, GETDATE())) AS VARCHAR(2))), 2)
                )
            ORDER BY d.whse, l.orderno;
        """,
        f"DL_SUPPLY_INV_{prev_month_yyyymm}": f"""
            SELECT
            {COMPANY_NAME} AS DistributorName,
            w.whse AS DistributorLocationNumber,
            CASE 
                WHEN LENGTH(d.zipcd) > 5 THEN SUBSTR(d.zipcd, 1, 5)
                ELSE d.zipcd
            END AS DistributorWarehousePostalCode,
            'US' AS DistributorWarehouseCountry,
            CASE WHEN vendprod <> '' THEN vendprod
                ELSE prod
                END AS ResideoSKUNumber,
            prod AS DistributorSKUNumber,
            DATE_FORMAT(CURRENT_DATE(), 'MM/dd/YYYY') AS InventoryDateSnapShotDate,
            CAST(GREATEST(qtyonhand - qtycommit - qtyreservd, 0) AS INT) AS QuantityAvailable,
            CAST(qtycommit + qtyreservd AS INT) AS QuantityAllocated,
            CAST(qtyonhand + qtycommit + qtyreservd AS INT) AS QuantityTotal,
            CAST(qtyonorder AS INT) AS OnOrderTotal,
            avgcost AS "Distributor'sAverageUnitCost", --MAY NOT BE RIGHT
            'USD' AS CurrencyBeingReported
            FROM default.icsw w
            LEFT JOIN icsd d on d.cono = 1 and d.whse = w.whse
            where 
            w.cono = 1
            AND w.prodline = 'RSIDEO'
            AND w.whse not like '%C'
            AND qtyonhand + qtycommit + qtyreservd > 0
            ORDER BY w.whse desc, w.prod desc
        """}
       
    # -------------------------
    #  EXPORT FUNCTION
    # -------------------------
    def export_to_excel(df, filename):
        # Keep all text as text (retain leading zeros)
        for col in df.columns:
            df[col] = df[col].astype(str)

        # Write to Excel
        excel_path = f"{export_folder}\\{filename}.xlsx"
        df.to_excel(excel_path, index=False)

        # Open workbook to format
        wb = load_workbook(excel_path)
        ws = wb.active

        # Header style
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Set body font and autofit columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Calibri", size=11)

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

        # Enable filters
        ws.auto_filter.ref = ws.dimensions

        wb.save(excel_path)
        wb.close()
        print(f"✅ Exported: {excel_path}")


    # -------------------------
    #  MAIN LOOP
    # -------------------------
    for name, sql in queries.items():
        df = pd.read_sql(sql, conn_remote)
        export_to_excel(df, name)

    conn_remote.close()
    print("✅ All reports complete.")