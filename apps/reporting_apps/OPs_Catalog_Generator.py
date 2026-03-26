# OPs Catalog Generator.py

def generate_ops_catalog():
    import sys
    from pathlib import Path


    # --- Ensure repo root is in Python path --
    repo_root = Path(__file__).resolve().parent
    while not (repo_root / "core").exists() and repo_root.parent != repo_root:
        repo_root = repo_root.parent

    sys.path.insert(0, str(repo_root))


    import jaydebeapi
    import sqlite3
    import pandas as pd
    from openpyxl import load_workbook
    from datetime import datetime
    import os
    from core.config import PATHS, JDBC




    # === CONFIG ===

    sqlite_path = PATHS["purchdata"]
    template_file = os.path.join(PATHS["catalogs"],"OPS template.xlsx")
    output_folder = PATHS["catalogs"]

    # === SQL SCRIPTS (hard-coded) ===
    external_query = """
    WITH RankedRows AS (
        SELECT
            whse,
            prod,
            baseprice,
            listprice,
            replcost,
            vendprod,
            qtyonhand,
            arpvendno,
            ROW_NUMBER() OVER (PARTITION BY prod ORDER BY CASE WHEN whse = '25' THEN 0 ELSE 1 END) AS RowNum
        FROM
            default.icsw
        WHERE
            cono = '1'
    ),
    ProductDetails AS (
        SELECT DISTINCT
            RankedRows.whse,
            RankedRows.prod,
            RankedRows.baseprice,
            RankedRows.listprice,
            RankedRows.replcost,
            RankedRows.vendprod,
            RankedRows.qtyonhand,
            RankedRows.arpvendno,
            icsp.descrip_1,
            icsp.prodcat,
            icsp.user10,
            icsp.brandcode,
            icsp.unitsell,
            icsp.unitstock
        FROM
            RankedRows
        JOIN
            default.icsp ON RankedRows.prod = icsp.prod
        WHERE
            RankedRows.RowNum = 1 AND icsp.cono = 1
    )
    SELECT
        ProductDetails.descrip_1 AS Product_Name,
        ProductDetails.descrip_1 AS Product_Description,
        CASE 
            WHEN ProductDetails.prodcat IN ('HCM', 'COEQ', 'COPT', 'HCEQ', 'HCPT','RF') THEN 'HVAC'
            ELSE 'Appliances'
        END AS Category_Main,
        CASE 
            WHEN ProductDetails.user10 = 'APDW' THEN 'Dishwasher'
            WHEN ProductDetails.user10 = 'APGB' THEN 'Disposals'
            WHEN ProductDetails.user10 = 'APIM' THEN 'Ice Makers'
            WHEN ProductDetails.user10 = 'APLA' THEN 'Laundry'
            WHEN ProductDetails.user10 = 'APMW' THEN 'Microwave'
            WHEN ProductDetails.user10 = 'APRH' THEN 'Range Hoods'
            WHEN ProductDetails.user10 = 'APRG' THEN 'Range/Oven'
            WHEN ProductDetails.user10 = 'APRF' THEN 'Refrigerator'
            WHEN ProductDetails.user10 = 'HVFU' THEN 'Furnace'
            WHEN ProductDetails.user10 = 'HVMS' THEN 'Air Conditioner'
            WHEN ProductDetails.user10 = 'HVPU' THEN 'Heating and Air Conditioning - Combination Units'
            WHEN ProductDetails.user10 = 'HVPT' THEN 'Heating and Air Conditioning - Combination Units'
            WHEN ProductDetails.user10 = 'HVSS' THEN 'Air Conditioner'
            WHEN ProductDetails.user10 = 'HVWU' THEN 'Air Conditioner'
            WHEN ProductDetails.user10 = 'HVCL' THEN 'Air Cleaners'
            WHEN ProductDetails.user10 = 'HVUV' THEN 'Air Cleaners'
            WHEN ProductDetails.user10 = 'TACH' THEN 'HVAC - Chemicals'
            WHEN ProductDetails.user10 = 'WHWH' THEN 'Water Heaters'
            WHEN ProductDetails.prodcat IN ('HCM', 'COEQ', 'COPT', 'HCEQ', 'HCPT','RF') THEN 'A/C & Heater Repair & Parts'
            ELSE 'Misc Appliance Parts' 
        END AS Category_Lower,
        tablecode.descrip AS Manufacturer,
        ProductDetails.prod AS Supplier_SKU,
        CASE 
            WHEN ProductDetails.unitsell <> '' THEN ProductDetails.unitsell
            ELSE ProductDetails.unitstock
        END AS UOM,
        CASE 
            WHEN ProductDetails.unitsell <> '' THEN ProductDetails.unitsell
            ELSE ProductDetails.unitstock
        END AS Package_Qty,
        ProductDetails.baseprice
    FROM
        ProductDetails
    LEFT JOIN (
        SELECT codeval, descrip 
        FROM default.sasta 
        WHERE cono = 1 AND codeiden = 'BC'
    ) tablecode ON ProductDetails.brandcode = tablecode.codeval
    WHERE
        ProductDetails.vendprod <> '!999!' AND ProductDetails.prod NOT LIKE '%+%'
            AND ProductDetails.prod NOT LIKE '%!%'
            AND ProductDetails.prod NOT LIKE '%#%'
            AND ProductDetails.prod NOT LIKE '%(%'
            AND ProductDetails.prod NOT LIKE '%)%'
    ORDER BY
        ProductDetails.prod;
    """

    local_query = """
    SELECT 
    '' as Update_Type,
    Product_Name,
    Product_Description,
    Category_Main,
    Category_Lower,
    CASE WHEN
        Category_Lower = 'Misc Appliance Parts' THEN '[1079]'
        ELSE '[338]'
        END AS Category_Code,
    Manufacturer,
    Supplier_SKU,
    UOM,
    Package_Qty,
    baseprice,
    CASE 
            WHEN d.imageFull1 IS NOT NULL THEN d.imageFull1
            WHEN d.image1 IS NOT NULL THEN d.image1
            WHEN s.StreamFlowImage1 IS NOT NULL THEN s.StreamFlowImage1
            ELSE ''
        END AS Image,
    '' as Replaced_By,
    '' as Prop_65
    from opsdata o
    LEFT JOIN ddsexport d on o.Supplier_SKU = d.ID
    LEFT JOIN sfimages s on o.Supplier_SKU = s."Item ID"
    """


    # === STEP 1: Run SQL script over JDBC and load results ===


    with jaydebeapi.connect(
        jclassname=JDBC["class"],
        url=JDBC["url"],
        driver_args=[JDBC["user"], JDBC["password"]],
        jars=JDBC["jar"]
    ) as conn_jdbc:
        df_external = pd.read_sql(external_query, conn_jdbc)

    # === STEP 2: Save results into SQLite (replace table) ===
    conn_sqlite = sqlite3.connect(sqlite_path)
    table_name = "opsdata"
    df_external.to_sql(table_name, conn_sqlite, if_exists="replace", index=False)

    # === STEP 3: Run 2nd query against SQLite ===

    df_local = pd.read_sql(local_query, conn_sqlite)
    conn_sqlite.close()

    # === STEP 4: Paste results into template starting at A20 ===
    wb = load_workbook(template_file)
    ws = wb.active  # or wb["SheetName"]

    start_row = 20
    start_col = 1  # column A

    for r_idx, row in enumerate(df_local.values, start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # === STEP 5: Save with date suffix ===
    today_str = datetime.today().strftime("%m%d%y")
    output_file = os.path.join(output_folder, f"OPS {today_str}.xlsx")
    wb.save(output_file)

    print(f"Done. File saved as {output_file}")

